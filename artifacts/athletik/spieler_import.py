"""Sicherer, speicherloser Import von Spieler-Stammdaten.

Dieses Modul enthält ausschließlich die Verarbeitung von Upload-Daten. Es schreibt
nie selbst in die Datenbank und übernimmt bewusst keine IDs, Rollen oder
Mandantenangaben aus einer Datei.
"""

from __future__ import annotations

import csv
import io
import re
import unicodedata
import zipfile
from datetime import date
from hashlib import sha256
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from database import altersklasse_vorschlag, parse_datum_safe
from utils.file_magic import validate_excel


MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_ROWS = 1_000
MAX_COLUMNS = 40
MAX_CELL_CHARS = 250

IMPORT_FIELDS = {
    "vorname": "Vorname *",
    "nachname": "Nachname *",
    "geburtsdatum": "Geburtsdatum *",
    "geschlecht": "Geschlecht",
    "hauptposition": "Hauptposition",
    "nebenposition": "Nebenposition",
    "spielbein": "Spielbein",
    "leistungsniveau": "Leistungsniveau",
    "mannschaft": "Mannschaft",
    "trainingsstatus": "Trainingsstatus",
}

OPTIONAL_DEFAULTS = {
    "geschlecht": "Männlich",
    "hauptposition": "Zentrales Mittelfeld",
    "nebenposition": "",
    "spielbein": "Rechts",
    "leistungsniveau": "Breitensport",
    "mannschaft": "",
    "trainingsstatus": "Uneingeschränktes Mannschaftstraining",
}

CANONICAL_POSITIONEN = [
    "Torwart", "Innenverteidiger", "Außenverteidiger (rechts)",
    "Außenverteidiger (links)", "Defensives Mittelfeld", "Zentrales Mittelfeld",
    "Offensives Mittelfeld", "Rechtes Mittelfeld", "Linkes Mittelfeld",
    "Rechter Flügel", "Linker Flügel", "Hängende Spitze", "Mittelstürmer",
]
CANONICAL_LEISTUNGSNIVEAUS = [
    "Breitensport", "Leistungssport", "Regionalkader", "Landeskader",
    "Bundeskader", "Profi",
]
CANONICAL_TRAININGSSTATUS = [
    "Uneingeschränktes Mannschaftstraining", "Angepasstes Mannschaftstraining",
    "Individuelles Training", "Trainingspause", "Externe Abklärung empfohlen",
    "Externe Freigabe dokumentiert",
]

_HEADER_ALIASES = {
    "vorname": {"vorname", "firstname", "first name", "rufname"},
    "nachname": {"nachname", "name", "familienname", "surname", "last name"},
    "geburtsdatum": {"geburtsdatum", "geburtstag", "birthdate", "date of birth", "dob"},
    "geschlecht": {"geschlecht", "gender", "sex"},
    "hauptposition": {"hauptposition", "position", "spielposition", "primary position"},
    "nebenposition": {"nebenposition", "zweite position", "secondary position"},
    "spielbein": {"spielbein", "fuss", "fuß", "foot", "starker fuss", "starker fuß"},
    "leistungsniveau": {"leistungsniveau", "niveau", "level"},
    "mannschaft": {"mannschaft", "team", "kader", "verein"},
    "trainingsstatus": {"trainingsstatus", "status", "training status"},
}

_GENDER_ALIASES = {
    "m": "Männlich", "mannlich": "Männlich", "männlich": "Männlich", "male": "Männlich",
    "w": "Weiblich", "weiblich": "Weiblich", "female": "Weiblich",
    "d": "Divers", "divers": "Divers", "diverse": "Divers",
}
_FOOT_ALIASES = {
    "rechts": "Rechts", "right": "Rechts", "r": "Rechts",
    "links": "Links", "left": "Links", "l": "Links",
    "beidfussig": "Beidfüßig", "beidfüßig": "Beidfüßig", "beidbeinig": "Beidfüßig",
    "both": "Beidfüßig",
}


def _plain(value: Any) -> str:
    """Konvertiert Zellwerte in begrenzten, formelfreien Text."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    value = str(value).strip()
    return value[:MAX_CELL_CHARS]


def _fold(value: Any) -> str:
    value = unicodedata.normalize("NFKD", _plain(value))
    value = "".join(char for char in value if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", value).strip().casefold()


def _unsafe_formula(value: str) -> bool:
    return bool(value.lstrip().startswith(("=", "+", "-", "@")))


def upload_fingerprint(file_bytes: bytes) -> str:
    """Nur ein Hash für Session-State; die Originaldatei wird nie abgelegt."""
    return sha256(file_bytes).hexdigest()


def auto_mapping(headers: list[str]) -> dict[str, str | None]:
    """Erstellt eine vorsichtige, eindeutige Header-Zuordnung."""
    available = {_fold(header): header for header in headers}
    mapping: dict[str, str | None] = {}
    used: set[str] = set()
    for field, aliases in _HEADER_ALIASES.items():
        match = next((available[alias] for alias in aliases if alias in available), None)
        if match and match not in used:
            mapping[field] = match
            used.add(match)
        else:
            mapping[field] = None
    return mapping


def _read_csv(file_bytes: bytes) -> tuple[list[str], list[dict[str, str]]]:
    if not file_bytes:
        raise ValueError("Die Datei ist leer.")
    if len(file_bytes) > MAX_UPLOAD_BYTES:
        raise ValueError("Die CSV-Datei ist größer als 10 MB.")

    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Die CSV-Datei hat keine unterstützte Zeichenkodierung.")

    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=";,\t|")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    headers = [str(header).strip() for header in (reader.fieldnames or []) if str(header).strip()]
    if not headers:
        raise ValueError("Die CSV-Datei enthält keine Kopfzeile.")
    if len(headers) > MAX_COLUMNS:
        raise ValueError(f"Die Datei enthält zu viele Spalten (maximal {MAX_COLUMNS}).")

    rows = []
    for number, row in enumerate(reader, start=2):
        if number > MAX_ROWS + 1:
            raise ValueError(f"Die Datei enthält mehr als {MAX_ROWS} Datenzeilen.")
        values = {header: _plain(row.get(header, "")) for header in headers}
        if any(values.values()):
            rows.append(values)
    return headers, rows


def _read_xlsx(file_bytes: bytes) -> tuple[list[str], list[dict[str, str]]]:
    ok, error = validate_excel(file_bytes, max_mb=10)
    if not ok:
        raise ValueError(error)

    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
            members = archive.namelist()
            if "[Content_Types].xml" not in members or "xl/workbook.xml" not in members:
                raise ValueError("Die Datei ist keine gültige XLSX-Arbeitsmappe.")
            uncompressed_size = sum(member.file_size for member in archive.infolist())
            if uncompressed_size > 50 * 1024 * 1024:
                raise ValueError("Die entpackte Arbeitsmappe ist zu groß.")
    except zipfile.BadZipFile as exc:
        raise ValueError("Die Excel-Datei ist beschädigt.") from exc

    try:
        workbook = load_workbook(
            filename=io.BytesIO(file_bytes),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        sheet = workbook[workbook.sheetnames[0]]
        raw_rows = sheet.iter_rows(values_only=False)
        header_cells = next(raw_rows, None)
        headers = [_plain(cell.value) for cell in (header_cells or [])]
        while headers and not headers[-1]:
            headers.pop()
        if not headers:
            raise ValueError("Die Excel-Datei enthält keine Kopfzeile.")
        if len(headers) > MAX_COLUMNS:
            raise ValueError(f"Die Datei enthält zu viele Spalten (maximal {MAX_COLUMNS}).")
        if len(set(_fold(header) for header in headers)) != len(headers):
            raise ValueError("Die Kopfzeile enthält doppelte Spaltennamen.")

        rows: list[dict[str, str]] = []
        for row_number, cells in enumerate(raw_rows, start=2):
            if row_number > MAX_ROWS + 1:
                raise ValueError(f"Die Datei enthält mehr als {MAX_ROWS} Datenzeilen.")
            values: dict[str, str] = {}
            for index, header in enumerate(headers):
                cell = cells[index] if index < len(cells) else None
                if cell is not None and cell.data_type == "f":
                    raise ValueError(f"Formeln sind nicht erlaubt (Zeile {row_number}).")
                values[header] = _plain(cell.value if cell is not None else "")
            if any(values.values()):
                rows.append(values)
        return headers, rows
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError("Die Excel-Datei konnte nicht gelesen werden.") from exc


def read_upload(file_bytes: bytes, filename: str) -> tuple[list[str], list[dict[str, str]]]:
    """Liest ausschließlich CSV oder XLSX in den Arbeitsspeicher."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _read_csv(file_bytes)
    if name.endswith(".xlsx"):
        return _read_xlsx(file_bytes)
    raise ValueError("Erlaubt sind nur CSV- und XLSX-Dateien.")


def validate_mapping(headers: list[str], mapping: dict[str, str | None]) -> list[str]:
    """Validiert die vom Nutzer gewählte Zuordnung vor dem Erstellen der Vorschau."""
    errors = []
    selected = [source for source in mapping.values() if source]
    if len(set(selected)) != len(selected):
        errors.append("Eine Datei-Spalte darf nur einem Importfeld zugeordnet werden.")
    invalid = [source for source in selected if source not in headers]
    if invalid:
        errors.append("Die gewählte Spaltenzuordnung passt nicht zur hochgeladenen Datei.")
    for required in ("vorname", "nachname", "geburtsdatum"):
        if not mapping.get(required):
            errors.append(f"Bitte ordne eine Spalte für {IMPORT_FIELDS[required]} zu.")
    return errors


def _normalise_choice(value: str, allowed: list[str], aliases: dict[str, str] | None = None) -> str | None:
    if not value:
        return ""
    folded = _fold(value)
    if aliases and folded in aliases:
        return aliases[folded]
    for option in allowed:
        if _fold(option) == folded:
            return option
    return None


def _normalise_date(value: str) -> str | None:
    parsed = parse_datum_safe(value)
    if parsed is None or parsed > date.today() or parsed.year < 1900:
        return None
    return parsed.strftime("%d.%m.%Y")


def _source_value(source: dict[str, str], mapped_header: str | None) -> str:
    return _plain(source.get(mapped_header, "")) if mapped_header else ""


def build_preview(
    source_rows: list[dict[str, str]],
    mapping: dict[str, str | None],
    existing_keys: set[tuple[str, str, str]],
    *,
    positionen: list[str],
    leistungsniveaus: list[str],
    trainingsstatus: list[str],
) -> list[dict[str, Any]]:
    """Erstellt Vorschauzeilen und bewertet sie rein serverseitig."""
    records = []
    for source_index, source in enumerate(source_rows, start=1):
        record: dict[str, Any] = {"_zeile": source_index, "Importieren": True}
        for field in IMPORT_FIELDS:
            record[field] = _source_value(source, mapping.get(field))
        records.append(record)
    return revalidate_preview(
        records,
        existing_keys,
        positionen=positionen,
        leistungsniveaus=leistungsniveaus,
        trainingsstatus=trainingsstatus,
    )


def revalidate_preview(
    records: list[dict[str, Any]],
    existing_keys: set[tuple[str, str, str]],
    *,
    positionen: list[str],
    leistungsniveaus: list[str],
    trainingsstatus: list[str],
) -> list[dict[str, Any]]:
    """Normalisiert editierte Vorschauwerte und ergänzt Status/Hinweise neu."""
    checked: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()
    for original in records:
        row = {"_zeile": int(original.get("_zeile", 0)), "Importieren": bool(original.get("Importieren", True))}
        notes: list[str] = []
        errors: list[str] = []

        for field in ("vorname", "nachname", "mannschaft"):
            value = _plain(original.get(field, ""))
            row[field] = value
            if value and _unsafe_formula(value):
                errors.append(f"{IMPORT_FIELDS[field]} enthält keine erlaubte Eingabe")

        for field in ("vorname", "nachname"):
            if not row[field]:
                errors.append(f"{IMPORT_FIELDS[field]} fehlt")

        raw_date = _plain(original.get("geburtsdatum", ""))
        if _unsafe_formula(raw_date):
            errors.append("Geburtsdatum enthält keine erlaubte Eingabe")
        date_value = _normalise_date(raw_date)
        if date_value is None:
            errors.append("Geburtsdatum ist ungültig oder liegt in der Zukunft")
            row["geburtsdatum"] = raw_date
            row["altersklasse"] = ""
        else:
            row["geburtsdatum"] = date_value
            row["altersklasse"] = altersklasse_vorschlag(date_value)

        choices = (
            ("geschlecht", ["Männlich", "Weiblich", "Divers"], _GENDER_ALIASES),
            ("hauptposition", positionen, None),
            ("nebenposition", [""] + positionen, None),
            ("spielbein", ["Rechts", "Links", "Beidfüßig"], _FOOT_ALIASES),
            ("leistungsniveau", leistungsniveaus, None),
            ("trainingsstatus", trainingsstatus, None),
        )
        for field, allowed, aliases in choices:
            raw = _plain(original.get(field, ""))
            if _unsafe_formula(raw):
                errors.append(f"{IMPORT_FIELDS[field]} enthält keine erlaubte Eingabe")
                row[field] = raw
                continue
            normalised = _normalise_choice(raw, allowed, aliases)
            if normalised is None:
                errors.append(f"{IMPORT_FIELDS[field]} ist nicht zulässig")
                row[field] = raw
            elif not normalised and field in OPTIONAL_DEFAULTS:
                row[field] = OPTIONAL_DEFAULTS[field]
                notes.append(f"{IMPORT_FIELDS[field]} ergänzt")
            else:
                row[field] = normalised

        duplicate_key = (
            _fold(row.get("vorname", "")),
            _fold(row.get("nachname", "")),
            row.get("geburtsdatum", ""),
        )
        if not errors and duplicate_key in existing_keys:
            notes.append("Bestehender Spieler im aktiven Mandanten – wird übersprungen")
        elif not errors and duplicate_key in seen:
            notes.append("Dublettenzeile innerhalb dieser Datei – wird übersprungen")
        elif not errors:
            seen.add(duplicate_key)

        if errors:
            row["status"] = "🔴 Fehler"
            row["hinweis"] = " · ".join(errors)
        elif notes:
            row["status"] = "🟡 Hinweis"
            row["hinweis"] = " · ".join(notes)
        else:
            row["status"] = "🟢 Bereit"
            row["hinweis"] = "Bereit zum Anlegen"
        checked.append(row)
    return checked


def import_candidates(records: list[dict[str, Any]]) -> list[dict[str, str]]:
    """Gibt nur ausdrücklich markierte, fehlerfreie Zeilen für den DB-Pfad zurück."""
    return [
        {field: _plain(row.get(field, "")) for field in IMPORT_FIELDS} | {
            "altersklasse": _plain(row.get("altersklasse", "")),
        }
        for row in records
        if row.get("Importieren") and row.get("status") != "🔴 Fehler"
    ]


def validate_write_row(raw: Any) -> dict[str, str] | None:
    """Validiert eine einzelne Schreibzeile ohne Vertrauen in die UI-Vorschau.

    Diese Funktion ist die serverseitige zweite Sicherheitsgrenze für den
    Datenbankpfad. Sie übernimmt keine unbekannten Felder und akzeptiert nur die
    kanonischen Auswahlwerte der Spielerstammdaten.
    """
    if not isinstance(raw, dict):
        return None
    record = {
        "_zeile": 1,
        "Importieren": True,
        **{field: raw.get(field, "") for field in IMPORT_FIELDS},
    }
    checked = revalidate_preview(
        [record],
        set(),
        positionen=CANONICAL_POSITIONEN,
        leistungsniveaus=CANONICAL_LEISTUNGSNIVEAUS,
        trainingsstatus=CANONICAL_TRAININGSSTATUS,
    )[0]
    if checked["status"] == "🔴 Fehler":
        return None
    return {
        field: _plain(checked.get(field, ""))
        for field in IMPORT_FIELDS
    } | {"altersklasse": _plain(checked.get("altersklasse", ""))}