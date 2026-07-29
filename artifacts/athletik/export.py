"""
export.py — Kader-Export als Excel-Datei
=========================================
Erstellt eine .xlsx-Datei mit zwei Tabellenblättern:
  1. Kader — alle Spieler mit Stammdaten und letzten Testwerten
  2. Verletzungshistorie — alle Verletzungen des gesamten Kaders
"""

import io
from datetime import date

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from database import (
    spieler_laden,
    spieler_by_id,
    verletzungen_laden,
    anthropometrie_letzter,
    anthropometrie_history,
    fms_letzter,
    fms_history,
    y_balance_letzter,
    y_balance_history,
    sprint_letzter,
    sprint_history,
    sprung_letzter,
    sprung_history,
    agilitaet_letzter,
    agilitaet_history,
    ausdauer_letzter,
    ausdauer_history,
    berechne_alter,
)
from analytics import risiko_score, risiko_label, athletik_score


# ─── Farben (dunkel, passend zum App-Theme) ───────────────────────────────────
_HDR_BG   = "1C2333"   # Kopfzeilen-Hintergrund (dunkelblau)
_HDR_FG   = "E6EDF3"   # Kopfzeilen-Schrift (hell)
_ALT_BG   = "161B27"   # Alternating row (dunkel)
_NORM_BG  = "0D1117"   # Normale Zeile
_ACCENT   = "238636"   # Grün (Athletik-Score gut)
_WARN     = "D29922"   # Gelb
_DANGER   = "F85149"   # Rot
_BORDER   = "30363D"   # Rahmenfarbe
_INJ_BG   = "1A1F2E"   # Verletzungsblatt Hintergrund
_INJ_HDR  = "6E40C9"   # Verletzungsblatt Header (lila)


def _thin_border():
    s = Side(style="thin", color=_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_style(cell, bg=_HDR_BG, fg=_HDR_FG, bold=True):
    cell.font      = Font(name="Calibri", bold=bold, color=fg, size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _thin_border()


def _data_style(cell, bg=_NORM_BG, fg="C9D1D9", bold=False, align="center"):
    cell.font      = Font(name="Calibri", bold=bold, color=fg, size=10)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _thin_border()


def _set_col_widths(ws, widths: dict):
    """widths: {col_letter: width}"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _val(d, key, fallback="—"):
    """Safe dict getter with fallback."""
    if not d:
        return fallback
    v = d.get(key)
    return v if v is not None else fallback


# ─── Blatt 1: Kader ──────────────────────────────────────────────────────────

_KADER_HEADERS = [
    # Stammdaten
    "Name", "Vorname", "Nachname", "Geburtsdatum", "Alter", "Geschlecht",
    "Altersklasse", "Hauptposition", "Nebenposition", "Spielbein",
    "Mannschaft", "Leistungsniveau", "Trainingsstatus",
    # Anthropometrie
    "Größe (cm)", "Gewicht (kg)", "BMI", "Körperfett (%)", "Muskelmasse (kg)",
    "Reifestatus",
    # Athletik Score & Risiko
    "Athletik Score", "Risiko",
    # FMS
    "FMS Score", "FMS Bewertung", "FMS Asymmetrie",
    # Y-Balance
    "YBal Composite R (%)", "YBal Composite L (%)", "YBal Asymmetrie",
    # Sprint
    "Sprint 5m (s)", "Sprint 10m (s)", "Sprint 20m (s)", "Sprint 30m (s)",
    "Sprint Beschl.-Index",
    # Sprung
    "CMJ beid. (cm)", "CMJ re (cm)", "CMJ li (cm)", "CMJ Asymmetrie (%)",
    "Squat Jump (cm)", "Drop Jump Höhe (cm)", "RSI", "Standweit (cm)",
    # Agilität
    "505 re (s)", "505 li (s)", "505 Asym. (%)", "5-10-5 (s)",
    "T-Test (s)", "Illinois (s)",
    # Ausdauer
    "Ausdauer Test-Typ", "Distanz (m)", "VO₂max", "HF max",
    # Datum letzter Tests
    "Datum FMS", "Datum Sprint", "Datum Sprung", "Datum Agilität", "Datum Ausdauer",
]


def _build_kader_sheet(ws):
    """Füllt Blatt 1 mit allen Spielern und ihren letzten Testwerten."""
    ws.title = "Kader"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # Kopfzeile
    for col_idx, header in enumerate(_KADER_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _header_style(cell)

    # Zeilen
    all_players = spieler_laden()
    for row_idx, p in enumerate(all_players, start=2):
        pid    = p["id"]
        anthr  = anthropometrie_letzter(pid)
        fms    = fms_letzter(pid)
        y      = y_balance_letzter(pid)
        sprint = sprint_letzter(pid)
        sprung = sprung_letzter(pid)
        agil   = agilitaet_letzter(pid)
        aus    = ausdauer_letzter(pid)
        verlet = []  # nur für Risiko-Score nötig — wir laden separat
        from database import verletzungen_laden as _vl
        verlet = _vl(pid)
        rs     = risiko_score(fms, y, verlet)
        _, level = risiko_label(rs)
        sc     = athletik_score(fms, y, sprint, sprung, agil, aus)
        alter  = berechne_alter(p.get("geburtsdatum")) or "—"

        risiko_de = {"hoch": "Hoch ⚠", "mittel": "Mittel", "gering": "Gering"}.get(level, level)

        row_data = [
            # Stammdaten
            p.get("name", "—"),
            p.get("vorname") or "—",
            p.get("nachname") or "—",
            p.get("geburtsdatum") or "—",
            alter,
            p.get("geschlecht") or "—",
            p.get("altersklasse") or "—",
            p.get("hauptposition") or p.get("position") or "—",
            p.get("nebenposition") or "—",
            p.get("spielbein") or "—",
            p.get("mannschaft") or "—",
            p.get("leistungsniveau") or "—",
            p.get("trainingsstatus") or "—",
            # Anthropometrie
            _val(anthr, "groesse"),
            _val(anthr, "gewicht"),
            _val(anthr, "bmi"),
            _val(anthr, "koerperfett"),
            _val(anthr, "muskelmasse"),
            _val(anthr, "reifestatus"),
            # Scores
            sc,
            risiko_de,
            # FMS
            _val(fms, "score"),
            _val(fms, "bewertung"),
            _val(fms, "asymmetrie"),
            # Y-Balance
            _val(y, "composite_rechts"),
            _val(y, "composite_links"),
            _val(y, "asymmetrie"),
            # Sprint
            _val(sprint, "beste_5m"),
            _val(sprint, "beste_10m"),
            _val(sprint, "beste_20m"),
            _val(sprint, "beste_30m"),
            _val(sprint, "beschl_index"),
            # Sprung
            _val(sprung, "cmj_beid"),
            _val(sprung, "cmj_rechts"),
            _val(sprung, "cmj_links"),
            _val(sprung, "cmj_asymmetrie"),
            _val(sprung, "squat_jump"),
            _val(sprung, "drop_jump_hoehe"),
            _val(sprung, "rsi"),
            _val(sprung, "standweit"),
            # Agilität
            _val(agil, "t505_r"),
            _val(agil, "t505_l"),
            _val(agil, "asym_505"),
            _val(agil, "t5_10_5"),
            _val(agil, "t_test"),
            _val(agil, "illinois"),
            # Ausdauer
            _val(aus, "test_typ"),
            _val(aus, "distanz_m"),
            _val(aus, "vo2max"),
            _val(aus, "hf_max"),
            # Testdaten
            _val(fms, "datum"),
            _val(sprint, "datum"),
            _val(sprung, "datum"),
            _val(agil, "datum"),
            _val(aus, "datum"),
        ]

        bg = _ALT_BG if row_idx % 2 == 0 else _NORM_BG
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Risiko-Spalte farbig hervorheben (Spalte 21)
            if col_idx == 21:  # Risiko
                if level == "hoch":
                    _data_style(cell, bg=_DANGER, fg="FFFFFF", bold=True)
                elif level == "mittel":
                    _data_style(cell, bg=_WARN, fg="0D1117", bold=True)
                else:
                    _data_style(cell, bg=_ACCENT, fg="FFFFFF", bold=True)
            elif col_idx == 20:  # Athletik Score
                _data_style(cell, bg=bg, fg="E6EDF3", bold=True, align="center")
            elif col_idx <= 13:  # Stammdaten linksbündig
                _data_style(cell, bg=bg, fg="C9D1D9", align="left")
            else:
                _data_style(cell, bg=bg, fg="C9D1D9", align="center")

    # Spaltenbreiten
    widths = {
        "A": 22, "B": 14, "C": 14, "D": 13, "E": 7,  "F": 11,
        "G": 14, "H": 20, "I": 16, "J": 12, "K": 18, "L": 14, "M": 22,
        "N": 11, "O": 12, "P": 8,  "Q": 12, "R": 14, "S": 14,
        "T": 14, "U": 12,
        "V": 11, "W": 16, "X": 14,
        "Y": 14, "Z": 14,
    }
    # Weitere Spalten (AA onward)
    extra = [12, 11, 11, 11, 13, 11, 11, 14, 13, 11, 11, 14, 14, 8, 8, 11, 10, 14,
             13, 14, 11, 13, 13, 13, 13]
    for i, w in enumerate(extra, start=27):
        widths[get_column_letter(i)] = w
    _set_col_widths(ws, widths)

    ws.row_dimensions[1].height = 36


# ─── Blatt 2: Verletzungshistorie ────────────────────────────────────────────

_INJ_HEADERS = [
    "Name", "Datum", "Verletzungsart", "Körperteil",
    "Schweregrad", "Ausfall (Tage)", "Notizen",
]


def _build_verletzung_sheet(ws):
    """Füllt Blatt 2 mit der Verletzungshistorie aller Spieler."""
    ws.title = "Verletzungshistorie"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    for col_idx, header in enumerate(_INJ_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _header_style(cell, bg=_INJ_HDR)

    all_players = spieler_laden()
    row_idx = 2
    for p in all_players:
        verlet = verletzungen_laden(p["id"])
        for v in verlet:
            bg = _ALT_BG if row_idx % 2 == 0 else _INJ_BG
            row_data = [
                p.get("name", "—"),
                v.get("datum") or "—",
                v.get("art") or "—",
                v.get("koerperteil") or "—",
                v.get("schwere") or "—",
                v.get("ausfall_tage") or "—",
                v.get("notizen") or "—",
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 5:  # Schweregrad farbig
                    schwere = str(value)
                    if "Schwer" in schwere:
                        _data_style(cell, bg=_DANGER, fg="FFFFFF", bold=True)
                    elif "Mittel" in schwere:
                        _data_style(cell, bg=_WARN, fg="0D1117", bold=True)
                    else:
                        _data_style(cell, bg=_ACCENT, fg="FFFFFF")
                elif col_idx == 7:  # Notizen linksbündig
                    _data_style(cell, bg=bg, fg="C9D1D9", align="left")
                else:
                    _data_style(cell, bg=bg, fg="C9D1D9", align="left" if col_idx == 1 else "center")
            row_idx += 1

    if row_idx == 2:
        # Keine Verletzungen vorhanden
        cell = ws.cell(row=2, column=1, value="Keine Verletzungen dokumentiert.")
        _data_style(cell, bg=_INJ_BG, fg="8B949E", align="left")

    _set_col_widths(ws, {
        "A": 22, "B": 13, "C": 22, "D": 18,
        "E": 22, "F": 15, "G": 40,
    })
    ws.row_dimensions[1].height = 30


# ─── Öffentliche API ─────────────────────────────────────────────────────────

def spieler_excel_bytes(spieler_id: int) -> bytes:
    """Einzelspieler-Export: Stammdaten, Testverlauf und Verletzungshistorie als Excel.

    Blatt 1  — Stammdaten & letzte Testwerte (Key-Value-Layout)
    Blatt 2  — Testverlauf aller Testmodule (chronologische Tabellen)
    Blatt 3  — Verletzungshistorie des Spielers
    """
    sp    = spieler_by_id(spieler_id)
    if not sp:
        raise ValueError(f"Spieler {spieler_id} nicht gefunden.")

    anthr  = anthropometrie_letzter(spieler_id)
    fms    = fms_letzter(spieler_id)
    y      = y_balance_letzter(spieler_id)
    sprint = sprint_letzter(spieler_id)
    sprung = sprung_letzter(spieler_id)
    agil   = agilitaet_letzter(spieler_id)
    aus    = ausdauer_letzter(spieler_id)
    verlet = verletzungen_laden(spieler_id)
    alter  = berechne_alter(sp.get("geburtsdatum")) or "—"
    rs     = risiko_score(fms, y, verlet)
    _, level = risiko_label(rs)
    sc     = athletik_score(fms, y, sprint, sprung, agil, aus)

    # History
    h_anthr  = anthropometrie_history(spieler_id)
    h_fms    = fms_history(spieler_id)
    h_y      = y_balance_history(spieler_id)
    h_sprint = sprint_history(spieler_id)
    h_sprung = sprung_history(spieler_id)
    h_agil   = agilitaet_history(spieler_id)
    h_aus    = ausdauer_history(spieler_id)

    wb = openpyxl.Workbook()

    # ─── Blatt 1: Stammdaten & letzte Testwerte ──────────────────────────────
    ws1 = wb.active
    ws1.title = "Stammdaten & Testwerte"
    ws1.sheet_view.showGridLines = False

    # Helper: KV-Zeile eintragen
    def _kv(row, key, val, key_bg=_HDR_BG):
        c_key = ws1.cell(row=row, column=1, value=key)
        c_key.font      = Font(name="Calibri", bold=True, color=_HDR_FG, size=10)
        c_key.fill      = PatternFill("solid", fgColor=key_bg)
        c_key.alignment = Alignment(horizontal="left", vertical="center")
        c_key.border    = _thin_border()
        c_val = ws1.cell(row=row, column=2, value=val)
        c_val.font      = Font(name="Calibri", color="C9D1D9", size=10)
        c_val.fill      = PatternFill("solid", fgColor=_NORM_BG)
        c_val.alignment = Alignment(horizontal="left", vertical="center")
        c_val.border    = _thin_border()

    def _section(row, title, bg="1C3A6B"):
        c = ws1.cell(row=row, column=1, value=title)
        c.font      = Font(name="Calibri", bold=True, color=_HDR_FG, size=11)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _thin_border()
        c2 = ws1.cell(row=row, column=2, value="")
        c2.fill   = PatternFill("solid", fgColor=bg)
        c2.border = _thin_border()
        return row + 1

    r = 1

    # Titel
    title_cell = ws1.cell(row=r, column=1,
                           value=f"Athletik-Bericht: {sp.get('name', '—')}")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=_HDR_FG)
    title_cell.fill = PatternFill("solid", fgColor="1046A0")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws1.row_dimensions[r].height = 28
    r += 1

    r = _section(r, "STAMMDATEN")
    for key, val in [
        ("Name",             sp.get("name") or "—"),
        ("Vorname",          sp.get("vorname") or "—"),
        ("Nachname",         sp.get("nachname") or "—"),
        ("Geburtsdatum",     sp.get("geburtsdatum") or "—"),
        ("Alter",            alter),
        ("Geschlecht",       sp.get("geschlecht") or "—"),
        ("Altersklasse",     sp.get("altersklasse") or "—"),
        ("Position (Haupt)", sp.get("hauptposition") or sp.get("position") or "—"),
        ("Position (Neben)", sp.get("nebenposition") or "—"),
        ("Spielbein",        sp.get("spielbein") or "—"),
        ("Mannschaft",       sp.get("mannschaft") or "—"),
        ("Leistungsniveau",  sp.get("leistungsniveau") or "—"),
        ("Trainingsstatus",  sp.get("trainingsstatus") or "—"),
    ]:
        _kv(r, key, val); r += 1

    r += 1
    r = _section(r, "ATHLETIK-KENNZAHLEN")
    sc_bg = _ACCENT if sc >= 75 else _WARN if sc >= 50 else _DANGER
    c_sc = ws1.cell(row=r, column=2, value=f"{sc}/100")
    c_sc.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    c_sc.fill = PatternFill("solid", fgColor=sc_bg)
    c_sc.alignment = Alignment(horizontal="left", vertical="center")
    c_sc.border = _thin_border()
    ws1.cell(row=r, column=1, value="Athletik-Score").font = Font(name="Calibri", bold=True, color=_HDR_FG, size=10)
    ws1.cell(row=r, column=1).fill = PatternFill("solid", fgColor=_HDR_BG)
    ws1.cell(row=r, column=1).border = _thin_border()
    r += 1
    level_de = {"hoch": "Handlungsbedarf hoch", "mittel": "Handlungsbedarf", "gering": "Unauffällig"}.get(level, level)
    _kv(r, "Athletik-Status", level_de); r += 1

    if anthr:
        r += 1
        r = _section(r, "ANTHROPOMETRIE (LETZTER WERT)")
        for key, field in [
            ("Datum",           "datum"),
            ("Größe (cm)",      "groesse"),
            ("Gewicht (kg)",    "gewicht"),
            ("BMI",             "bmi"),
            ("BMI-Kategorie",   "bmi_kategorie"),
            ("Körperfett (%)",  "koerperfett"),
            ("Muskelmasse (kg)","muskelmasse"),
            ("Reifestatus",     "reifestatus"),
        ]:
            _kv(r, key, _val(anthr, field)); r += 1

    for mod_name, row_data, fields in [
        ("FMS (LETZTER WERT)", fms, [
            ("Datum","datum"),("Gesamtscore","score"),("Bewertung","bewertung"),
            ("Asymmetrie","asymmetrie"),("Schwerpunkt","schwerpunkt"),
        ]),
        ("Y-BALANCE (LETZTER WERT)", y, [
            ("Datum","datum"),("Composite Rechts (%)","composite_rechts"),
            ("Composite Links (%)","composite_links"),
            ("Asymmetrie","asymmetrie"),("Schwerpunkt","schwerpunkt"),
        ]),
        ("SPRINT (LETZTER WERT)", sprint, [
            ("Datum","datum"),("Beste 5m (s)","beste_5m"),("Beste 10m (s)","beste_10m"),
            ("Beste 20m (s)","beste_20m"),("Beste 30m (s)","beste_30m"),
            ("Bewertung 10m","bewertung_10m"),("Bewertung 30m","bewertung_30m"),
        ]),
        ("SPRUNG (LETZTER WERT)", sprung, [
            ("Datum","datum"),("CMJ beidbeinig (cm)","cmj_beid"),
            ("CMJ rechts (cm)","cmj_rechts"),("CMJ links (cm)","cmj_links"),
            ("CMJ-Asymmetrie (%)","cmj_asymmetrie"),
            ("Squat Jump (cm)","squat_jump"),("RSI","rsi"),("Standweit (cm)","standweit"),
            ("Bewertung CMJ","bewertung_cmj"),
        ]),
        ("AGILITÄT (LETZTER WERT)", agil, [
            ("Datum","datum"),("505-Test rechts (s)","t505_r"),
            ("505-Test links (s)","t505_l"),("Asymmetrie 505 (%)","asym_505"),
            ("T-Test (s)","t_test"),("5-10-5 Shuttle (s)","t5_10_5"),
            ("Illinois (s)","illinois"),("Bewertung T-Test","bew_t_test"),
        ]),
        ("AUSDAUER / YO-YO (LETZTER WERT)", aus, [
            ("Datum","datum"),("Test-Level","test_typ"),
            ("Distanz (m)","distanz_m"),("VO2max","vo2max"),
            ("HF max (bpm)","hf_max"),("Bewertung","bewertung"),
        ]),
    ]:
        if row_data:
            r += 1
            r = _section(r, mod_name)
            for label, field in fields:
                _kv(r, label, _val(row_data, field)); r += 1

    ws1.column_dimensions["A"].width = 26
    ws1.column_dimensions["B"].width = 36

    # ─── Blatt 2: Testverlauf ─────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Testverlauf")
    ws2.sheet_view.showGridLines = False

    def _mod_history(ws, start_row, mod_title, headers, rows, fields, hdr_bg=_HDR_BG):
        """Schreibt einen Testmodul-Abschnitt (Header + Datenzeilen) ins Sheet."""
        if not rows:
            return start_row
        # Section header
        c = ws.cell(row=start_row, column=1, value=mod_title)
        c.font = Font(name="Calibri", bold=True, color=_HDR_FG, size=11)
        c.fill = PatternFill("solid", fgColor=hdr_bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=len(headers))
        ws.row_dimensions[start_row].height = 20
        r = start_row + 1
        # Column headers
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=ci, value=h)
            _header_style(cell)
        r += 1
        # Data rows
        for ri, row in enumerate(rows):
            bg = _ALT_BG if ri % 2 == 0 else _NORM_BG
            for ci, field in enumerate(fields, start=1):
                v = row.get(field) if isinstance(field, str) else field(row)
                cell = ws.cell(row=r, column=ci, value=v)
                _data_style(cell, bg=bg, align="left" if ci == 1 else "center")
            r += 1
        return r + 1   # blank row between sections

    r2 = 1
    title2 = ws2.cell(row=r2, column=1,
                       value=f"Testverlauf: {sp.get('name', '—')}")
    title2.font = Font(name="Calibri", bold=True, size=14, color=_HDR_FG)
    title2.fill = PatternFill("solid", fgColor="1046A0")
    title2.alignment = Alignment(horizontal="left", vertical="center")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws2.row_dimensions[1].height = 28
    r2 = 3

    r2 = _mod_history(ws2, r2, "FMS - FUNCTIONAL MOVEMENT SCREEN",
        ["Datum","Score /21","Bewertung","Asymmetrie","Schwerpunkt"],
        h_fms,
        ["datum","score","bewertung","asymmetrie","schwerpunkt"],
    )
    r2 = _mod_history(ws2, r2, "ANTHROPOMETRIE",
        ["Datum","Größe","Gewicht","KF %","Muskel kg","BMI","BMI-Kat.","Reifestatus"],
        h_anthr,
        ["datum","groesse","gewicht","koerperfett","muskelmasse","bmi","bmi_kategorie","reifestatus"],
        hdr_bg="1D4A3A",
    )
    r2 = _mod_history(ws2, r2, "Y-BALANCE",
        ["Datum","Comp. Rechts %","Comp. Links %","Asymmetrie","Schwerpunkt"],
        h_y,
        ["datum","composite_rechts","composite_links","asymmetrie","schwerpunkt"],
        hdr_bg="2D3F1D",
    )
    r2 = _mod_history(ws2, r2, "SPRINT-DIAGNOSTIK",
        ["Datum","Beste 5m (s)","Beste 10m (s)","Beste 20m (s)","Beste 30m (s)","Bew. 10m"],
        h_sprint,
        ["datum","beste_5m","beste_10m","beste_20m","beste_30m","bewertung_10m"],
        hdr_bg="3D2D1D",
    )
    r2 = _mod_history(ws2, r2, "SPRUNG-DIAGNOSTIK",
        ["Datum","CMJ beid. cm","CMJ R cm","CMJ L cm","Asym. %","SJ cm","RSI","Standweit cm"],
        h_sprung,
        ["datum","cmj_beid","cmj_rechts","cmj_links","cmj_asymmetrie","squat_jump","rsi","standweit"],
        hdr_bg="3D1D2D",
    )
    r2 = _mod_history(ws2, r2, "AGILITÄT",
        ["Datum","505 R (s)","505 L (s)","Asym. %","T-Test (s)","5-10-5 (s)","Illinois (s)"],
        h_agil,
        ["datum","t505_r","t505_l","asym_505","t_test","t5_10_5","illinois"],
        hdr_bg="1D2D3D",
    )
    r2 = _mod_history(ws2, r2, "AUSDAUER / YO-YO",
        ["Datum","Level","Distanz (m)","VO2max","HF max","Bewertung"],
        h_aus,
        ["datum","test_typ","distanz_m","vo2max","hf_max","bewertung"],
        hdr_bg="2D1D3D",
    )

    for col in range(1, 9):
        ws2.column_dimensions[get_column_letter(col)].width = [14, 13, 13, 13, 13, 13, 13, 14][col - 1]

    # ─── Blatt 3: Verletzungshistorie ─────────────────────────────────────────
    ws3 = wb.create_sheet(title="Verletzungshistorie")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A2"

    v_headers = ["Datum","Verletzungsart","Körperteil","Schweregrad","Ausfall (Tage)","Notizen"]
    for ci, h in enumerate(v_headers, start=1):
        cell = ws3.cell(row=1, column=ci, value=h)
        _header_style(cell, bg=_INJ_HDR)

    if not verlet:
        c = ws3.cell(row=2, column=1, value="Keine Verletzungen dokumentiert.")
        _data_style(c, bg=_INJ_BG, fg="8B949E", align="left")
    else:
        for ri, v in enumerate(verlet, start=2):
            bg = _ALT_BG if ri % 2 == 0 else _INJ_BG
            for ci, (field, align) in enumerate([
                ("datum", "center"),("art","left"),("koerperteil","left"),
                ("schwere","center"),("ausfall_tage","center"),("notizen","left"),
            ], start=1):
                val = v.get(field) or "—"
                cell = ws3.cell(row=ri, column=ci, value=val)
                if ci == 4:  # Schweregrad farbig
                    schwere = str(val)
                    if "Schwer" in schwere:
                        _data_style(cell, bg=_DANGER, fg="FFFFFF", bold=True, align="center")
                    elif "Mittel" in schwere:
                        _data_style(cell, bg=_WARN, fg="0D1117", bold=True, align="center")
                    else:
                        _data_style(cell, bg=_ACCENT, fg="FFFFFF", align="center")
                else:
                    _data_style(cell, bg=bg, fg="C9D1D9", align=align)

    _set_col_widths(ws3, {"A":13,"B":22,"C":18,"D":22,"E":15,"F":42})
    ws3.row_dimensions[1].height = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def kader_excel_bytes() -> bytes:
    """Gibt den kompletten Excel-Export als Bytes-Objekt zurück.
    Kann direkt an st.download_button(data=...) übergeben werden."""
    wb = openpyxl.Workbook()
    ws_kader = wb.active
    _build_kader_sheet(ws_kader)

    ws_inj = wb.create_sheet()
    _build_verletzung_sheet(ws_inj)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
